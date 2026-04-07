"""One-off: dump World_map.xlsx grid + legend for world_map_data.js generation."""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "World_map.xlsx"


def rgb_to_hex(cell):
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    if rgb is None:
        return None
    if isinstance(rgb, str):
        s = rgb.upper()
        if s.startswith("FF") and len(s) == 8:
            s = s[2:]
        if len(s) == 6:
            return "#" + s
    return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb.active
    print("sheet", ws.title, "rows", ws.max_row, "cols", ws.max_column)

    # Legend from row 104
    print("\n=== Legend (row 104+) ===")
    for r in range(104, min(130, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(8, ws.max_column + 1)):
            cell = ws.cell(r, c)
            row_vals.append(cell.value)
            hx = rgb_to_hex(cell)
            if hx and c <= 3:
                pass
        if any(x is not None and str(x).strip() for x in row_vals):
            print(r, row_vals)

    # Map grid: assume row1 = x labels, col A = y labels
    print("\n=== Row 1 (x coords) ===")
    xs = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        xs.append(v)
    print("first 20", xs[:20], "len", len(xs))

    print("\n=== Col A (y coords) ===")
    ys = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        ys.append(v)
    print("first 10", ys[:10], "len", len(ys))

    # Sample colors from data area row 2 col 2
    print("\n=== Sample colors r2-5 c2-8 ===")
    for r in range(2, 7):
        line = []
        for c in range(2, 10):
            hx = rgb_to_hex(ws.cell(r, c))
            line.append(hx or "?")
        print(r, line)

    # Unique colors in grid
    uniq = set()
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            hx = rgb_to_hex(ws.cell(r, c))
            if hx:
                uniq.add(hx)
    print("\nunique hex count", len(uniq))
    print(sorted(uniq)[:40])


if __name__ == "__main__":
    main()
